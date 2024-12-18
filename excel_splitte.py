import os
import openpyxl
import xlrd
from openpyxl.utils import get_column_letter

def split_excel_with_debug(input_file_path, output_directory_path):
    try:
        if input_file_path.endswith('.xls'):
            workbook = xlrd.open_workbook(input_file_path)
            is_xls = True
        elif input_file_path.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(input_file_path, read_only=True)
            is_xls = False
        else:
            print(f"Unsupported file format: {input_file_path}")
            return

        for sheet_name in (workbook.sheet_names() if is_xls else workbook.sheetnames):
            if "インデックス" in sheet_name:
                continue

            # 元ファイル名を含む出力名
            safe_sheet_name = sheet_name.replace(" ", "_").replace("/", "_")
            base_name = os.path.splitext(os.path.basename(input_file_path))[0]
            output_file_name = f"{base_name}_{safe_sheet_name}.xlsx"
            output_file_path = os.path.join(output_directory_path, output_file_name)
            print(f"Processing sheet: {sheet_name} -> {output_file_name}")

            try:
                new_workbook = openpyxl.Workbook()
                new_sheet = new_workbook.active

                if is_xls:
                    sheet = workbook.sheet_by_name(sheet_name)
                    for rx in range(sheet.nrows):
                        new_sheet.append(sheet.row_values(rx))
                    print(f"  Rows processed: {sheet.nrows}")
                else:
                    sheet = workbook[sheet_name]
                    row_count = 0
                    for row in sheet.iter_rows():
                        new_sheet.append([cell.value for cell in row])
                        row_count += 1
                    print(f"  Rows processed: {row_count}")

                new_workbook.save(output_file_path)
            except Exception as e:
                print(f"  Error processing sheet {sheet_name}: {e}")

    except Exception as e:
        print(f"Error processing file {input_file_path}: {e}")

def process_katashiki_with_debug(input_directory_path, output_directory_path):
    if not os.path.exists(output_directory_path):
        os.makedirs(output_directory_path)

    for filename in os.listdir(input_directory_path):
        if filename.endswith(('.xlsx', '.xls')):
            file_path = os.path.join(input_directory_path, filename)
            print(f"Processing file: {filename}")
            split_excel_with_debug(file_path, output_directory_path)

input_directory = "katashiki"
output_directory = "output"
process_katashiki_with_debug(input_directory, output_directory)
